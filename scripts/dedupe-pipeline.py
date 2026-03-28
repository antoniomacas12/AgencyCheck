"""
AgencyCheck — Import, Dedupe & Canonical Agency Pipeline v2
Phases 1-6: staging → normalize → domain-dedupe → fuzzy-flag → canonical → enrich

Key improvements over v1:
  - Better canonical name selection (shortest meaningful brand name)
  - Proper multi-city merging across all merged records
  - Housing/transport enum uppercased for Prisma (YES/NO/UNKNOWN)
  - OTTO Work Force I/II/III/IV → "OTTO Work Force"
  - Franchise/sub-brand groups use domain prefix as canonical hint
  - Idempotent slug generation (dedup collisions by appending city)
  - Richer verification status detection from notes
  - Housing notes keyword detection even for "unknown" rows

Output JSON files (scripts/data/):
  00_summary.json          — pipeline stats
  01_staging.json          — raw import with normalization fields
  02_fuzzy_candidates.json — all fuzzy pairs >= 0.80
  03_canonical_agencies.json — 1 record per real brand
  04_duplicate_map.json    — source row → canonical mapping
  05_branch_cities.json    — agencies with multiple known cities
  06_aliases.json          — canonical → [alias names]
  07_review_flags.json     — fuzzy pairs NOT auto-merged (for manual review)
"""

import json, re, unicodedata, sys
from difflib import SequenceMatcher
from urllib.parse import urlparse
from pathlib import Path
from collections import defaultdict
from datetime import date

# ─── install openpyxl if needed ───────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

XLSX_PATH = Path("/sessions/lucid-confident-hopper/mnt/uploads/netherlands_employment_agencies_research_expanded.xlsx")
OUT_DIR   = Path("/sessions/lucid-confident-hopper/mnt/outputs/agencycheck/scripts/data")
OUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = date.today().isoformat()   # "2026-03-11"

# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def clean(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "") else s

def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")

def root_domain(url):
    if not url: return ""
    raw = url.strip().lower()
    if not raw.startswith("http"):
        raw = "https://" + raw
    try:
        host = urlparse(raw).netloc or urlparse(raw).path.split("/")[0]
        host = re.sub(r"^www\.", "", host)
        return host.strip("/").split(":")[0]
    except Exception:
        return url.lower().strip()

def normalize_housing(val):
    """Normalize housing/transport value to YES/NO/UNKNOWN (Prisma enum)."""
    v = str(val).strip().lower()
    if v in ("yes", "true", "1", "ja", "provided", "available"):
        return "YES"
    if v in ("no", "false", "0", "nee", "not provided"):
        return "NO"
    return "UNKNOWN"

def normalize_name(name):
    """Canonical name for dedup comparison — strip noise before matching."""
    n = name.lower().strip()
    n = unicodedata.normalize("NFKD", n).encode("ascii", "ignore").decode()
    n = re.sub(r"[''`]", "", n)
    n = re.sub(r"[^\w\s]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    # Remove generic Dutch legal / structural suffixes
    noise = [
        r"\bb\.?v\.?\b", r"\bn\.?v\.?\b", r"\bnl\b", r"\bnederland\b",
        r"\buitzendbureau\b", r"\buitzenden\b", r"\bpayroll\b",
        r"\bworkforce\b", r"\bwork\s*force\b",
        r"\bpersoneelsdiensten\b",
        r"\bpersoneelsbureau\b",
        r"\bdetachering\b",
        r"\bpersoneel\b",
        r"\bservices?\b",
        r"\bcontracting\b",
        r"\bfacility\b",
        r"\bhospitality\b",
        r"\bpubliek\b",
        r"\boverheid\b",
        r"\bplus\b",
        r"\bzorg\b", r"\bwelzijn\b",
        r"\bstaff\b",
        r"\bformule\b",
        r"\btempsolution\b",
        # Roman numerals at the end
        r"\b(i{1,3}|iv|vi?|vii?|viii?|ix|x)\b",
    ]
    for pat in noise:
        n = re.sub(pat, " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n

def fuzzy(a, b):
    return SequenceMatcher(None, a, b).ratio()

def common_word_prefix(names):
    """
    Find the longest shared word-level prefix across all names.
    e.g. ["Abeos Bouwflex", "Abeos Agrarisch Bedrijf Service", "Abeos Bouwservice"]
         → "Abeos"
    e.g. ["Pro Industry Food", "Pro Industry Chemical", "Pro Industry Pharma"]
         → "Pro Industry"
    Returns None if no shared prefix (first word at minimum must match).
    """
    if not names:
        return None
    word_lists = [n.split() for n in names]
    min_len = min(len(wl) for wl in word_lists)
    prefix_words = []
    for i in range(min_len):
        word = word_lists[0][i].lower()
        if all(wl[i].lower() == word for wl in word_lists):
            prefix_words.append(word_lists[0][i])   # preserve original capitalisation
        else:
            break
    return " ".join(prefix_words) if prefix_words else None


def canonical_brand_name(names):
    """
    From a list of names that are all variants of the same brand, return the
    best canonical brand-level name.

    Strategy (in order):
    1. Try common_word_prefix — covers "Abeos *", "Pro Industry *", "Aethon *"
    2. Try stripping well-known trailing noise words (roman numerals, generic suffixes)
    3. Fall back to the shortest name

    Minimum length guard: result must be >= 3 chars.
    """
    # --- Strategy 1: Common word prefix ---
    prefix = common_word_prefix(names)
    if prefix and len(prefix) >= 3:
        return prefix

    # --- Strategy 2: Strip trailing noise from each name, pick most common result ---
    TRAILING_NOISE = re.compile(
        r"\s+(i{1,3}|iv|vi?|vii?|viii?|ix|x"
        r"|payroll|overheid|plus|uitzendbureau|uitzenden"
        r"|publiek|zorg|welzijn|services?|contracting|logistics?|facility"
        r"|hospitality|bouw|techniek|productie|detachering|formule"
        r"|tempsolution|openbaar\s*vervoer|transport"
        r"|kids|publica|zorgt|chemical|pharma|food|metal|logistic"
        r"|administratief|agrarisch|zakelijke\s*dienstverlening"
        r"|bouwservice|bouwflex|cross\s*border"
        r"|business\s*services?|werk\s*detachering"
        r")\s*$",
        re.IGNORECASE
    )
    stripped = []
    for n in names:
        s = n
        for _ in range(6):          # up to 6 passes to strip multiple suffixes
            s2 = TRAILING_NOISE.sub("", s).strip()
            if s2 == s:
                break
            s = s2
        if len(s) >= 3:
            stripped.append(s)

    if stripped:
        # Count how many original names start with each candidate (case-insensitive)
        counts = defaultdict(int)
        for s in stripped:
            for n in names:
                if n.lower().startswith(s.lower()):
                    counts[s] += 1
        best = max(stripped, key=lambda s: (counts[s], -len(s)))
        if len(best) >= 3:
            return best

    # --- Strategy 3: Return shortest name ---
    return min(names, key=len)

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1 — LOAD & STAGING
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
raw_rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

print(f"[Phase 1] Loaded {len(raw_rows)} raw rows, columns: {headers}")

staging = []
for i, row in enumerate(raw_rows):
    staging.append({
        "raw_id":           i,
        "raw_agency_name":  clean(row.get("agency_name")),
        "raw_website":      clean(row.get("website")),
        "raw_city":         clean(row.get("city")),
        "raw_housing":      normalize_housing(row.get("housing") or "unknown"),
        "raw_transport":    normalize_housing(row.get("transport") or "unknown"),
        "raw_job_types":    clean(row.get("job_types")) or "unknown",
        "raw_salary_range": clean(row.get("salary_range")) or "unknown",
        "raw_notes":        clean(row.get("notes")),
        "raw_source":       clean(row.get("source")),
    })

(OUT_DIR / "01_staging.json").write_text(json.dumps(staging, indent=2, ensure_ascii=False))
print(f"[Phase 1] Staged {len(staging)} rows")

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2A — NORMALIZE + ENRICH FIELDS
# ─────────────────────────────────────────────────────────────────────────────

HOUSING_KEYWORDS = ["housing", "huisvesting", "accommod", "woning", "slaap",
                    "verblijf", "logies", "kamer", "woonruimte"]
TRANSPORT_KEYWORDS = ["transport", "bus ", "busvervoer", "pendel", "shuttle",
                      "vervoer", "pendeldienst"]

for r in staging:
    r["norm_name"]   = normalize_name(r["raw_agency_name"])
    r["root_domain"] = root_domain(r["raw_website"])
    r["slug"]        = slugify(r["raw_agency_name"])
    r["city_clean"]  = r["raw_city"].strip().title() if r["raw_city"] else "unknown"

    # Detect housing/transport from notes even if field = UNKNOWN
    notes_lower = r["raw_notes"].lower() if r["raw_notes"] else ""
    if r["raw_housing"] == "UNKNOWN":
        if any(kw in notes_lower for kw in HOUSING_KEYWORDS):
            r["raw_housing"] = "YES"
            r["housing_from_notes"] = True
    if r["raw_transport"] == "UNKNOWN":
        if any(kw in notes_lower for kw in TRANSPORT_KEYWORDS):
            r["raw_transport"] = "YES"
            r["transport_from_notes"] = True

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2B — DOMAIN GROUPING (primary dedup signal)
# ─────────────────────────────────────────────────────────────────────────────

domain_groups = defaultdict(list)
for r in staging:
    if r["root_domain"]:
        domain_groups[r["root_domain"]].append(r["raw_id"])

domain_dupes = {d: ids for d, ids in domain_groups.items() if len(ids) > 1}
print(f"\n[Phase 2B] Domain duplicate groups: {len(domain_dupes)}")
for d, ids in sorted(domain_dupes.items(), key=lambda x: -len(x[1])):
    names = [staging[i]["raw_agency_name"] for i in ids]
    print(f"  {d} ({len(ids)}): {names[:4]}{'…' if len(names)>4 else ''}")

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2C — FUZZY NAME MATCHING
#   Auto-merge only when:
#   - score >= 0.95  (very close names)
#   - OR same domain (already in domain_dupes, handled above)
#   Everything else is flagged for review only
# ─────────────────────────────────────────────────────────────────────────────

already_in_domain_dupe = set()
for ids in domain_dupes.values():
    already_in_domain_dupe.update(ids)

rows_for_fuzzy = [r for r in staging if r["raw_id"] not in already_in_domain_dupe]
fuzzy_all = []
fuzzy_auto_merge = []

for i, ra in enumerate(rows_for_fuzzy):
    for rb in rows_for_fuzzy[i+1:]:
        na, nb = ra["norm_name"], rb["norm_name"]
        if not na or not nb or len(na) < 4 or len(nb) < 4:
            continue
        score = fuzzy(na, nb)
        if score < 0.80:
            continue
        same_domain = bool(ra["root_domain"] and ra["root_domain"] == rb["root_domain"])
        auto = score >= 0.95   # strict — only very close
        entry = {
            "id_a": ra["raw_id"], "name_a": ra["raw_agency_name"],
            "id_b": rb["raw_id"], "name_b": rb["raw_agency_name"],
            "norm_a": na, "norm_b": nb,
            "score": round(score, 3),
            "same_domain": same_domain,
            "auto_merge": auto,
        }
        fuzzy_all.append(entry)
        if auto:
            fuzzy_auto_merge.append(entry)

fuzzy_all.sort(key=lambda x: -x["score"])
(OUT_DIR / "02_fuzzy_candidates.json").write_text(
    json.dumps(fuzzy_all, indent=2, ensure_ascii=False))

print(f"\n[Phase 2C] Fuzzy pairs total: {len(fuzzy_all)}  auto-merge: {len(fuzzy_auto_merge)}")
print("  Pairs flagged for review:")
for p in fuzzy_all:
    if not p["auto_merge"]:
        print(f"    [score:{p['score']}] {p['name_a']!r} ↔ {p['name_b']!r}")

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3 — BUILD MERGE GROUPS
# ─────────────────────────────────────────────────────────────────────────────

merge_groups = {}
gid = 0
for domain, ids in domain_dupes.items():
    merge_groups[gid] = {"ids": set(ids), "reason": "same_domain", "domain": domain}
    gid += 1

def find_group(raw_id):
    for g, info in merge_groups.items():
        if raw_id in info["ids"]:
            return g
    return None

for pair in fuzzy_auto_merge:
    ga = find_group(pair["id_a"])
    gb = find_group(pair["id_b"])
    if ga is None and gb is None:
        merge_groups[gid] = {
            "ids": {pair["id_a"], pair["id_b"]},
            "reason": "fuzzy_name",
            "score":  pair["score"]
        }
        gid += 1
    elif ga is not None and gb is None:
        merge_groups[ga]["ids"].add(pair["id_b"])
    elif ga is None and gb is not None:
        merge_groups[gb]["ids"].add(pair["id_a"])
    elif ga != gb:
        merge_groups[ga]["ids"].update(merge_groups[gb]["ids"])
        del merge_groups[gb]

print(f"\n[Phase 3] Merge groups: {len(merge_groups)}")

id_to_group = {}
for g, info in merge_groups.items():
    for raw_id in info["ids"]:
        id_to_group[raw_id] = g

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3B — PRODUCE CANONICAL RECORDS
# ─────────────────────────────────────────────────────────────────────────────

def pick_best(*vals):
    """Return first non-empty, non-unknown value."""
    for v in vals:
        if v and str(v).strip().lower() not in ("unknown", ""):
            return v
    return "unknown"

def pick_enum_best(*vals):
    """For YES/NO/UNKNOWN: prefer YES > NO > UNKNOWN."""
    for v in vals:
        if v == "YES":
            return "YES"
    for v in vals:
        if v == "NO":
            return "NO"
    return "UNKNOWN"

def merge_job_types(*vals):
    items = set()
    for v in vals:
        if v and v not in ("unknown", ""):
            for part in re.split(r"[,/]", v):
                p = part.strip().lower()
                if p and p != "unknown":
                    items.add(p)
    return " / ".join(sorted(items)) if items else None

def merge_notes(*notes):
    seen, unique = set(), []
    for n in notes:
        if n and n not in ("unknown", "") and n not in seen:
            seen.add(n)
            unique.append(n)
    return " | ".join(unique) or None

def normalize_website(w):
    if not w: return None
    w = w.strip().lower()
    if not w.startswith("http"):
        w = "https://" + w
    return w.rstrip("/")

def build_web_pages(website, housing):
    if not website: return {}
    base = normalize_website(website)
    pages = {
        "homepage": base,
        "jobs":     base + "/vacatures",
        "contact":  base + "/contact",
    }
    if housing == "YES":
        pages["housing"] = base + "/huisvesting"
    return pages

def detect_verification_status(notes, val, from_notes=False):
    """Detect housing/transport verification status from context."""
    if val != "YES":
        return "unknown"
    if from_notes:
        return "worker_reported"
    notes_lower = (notes or "").lower()
    if any(kw in notes_lower for kw in ["official site", "website says", "official website", "official page"]):
        return "verified"
    if any(kw in notes_lower for kw in ["reported", "workers say", "workers report"]):
        return "worker_reported"
    return "verified"   # default: if dataset says YES, treat as verified

def detect_source_type(source_url):
    url = (source_url or "").lower()
    if "abu.nl" in url:
        return "ABU_REGISTER"
    if "sncu.nl" in url:
        return "SNCU_REGISTER"
    if url:
        return "OFFICIAL_WEBSITE"
    return "UNKNOWN"

canonical_agencies = []
dup_map_entries = []
processed = set()
canon_id = 0

def make_canonical(rows, merge_reason):
    global canon_id

    def completeness(r):
        score = 0
        if r["raw_housing"] == "YES": score += 4
        if r["raw_transport"] == "YES": score += 3
        if r["raw_job_types"] != "unknown": score += 2
        if r["raw_notes"]: score += 1
        score -= len(r["raw_agency_name"]) * 0.01  # prefer shorter brand names
        return score

    rows_sorted = sorted(rows, key=completeness, reverse=True)

    # --- Canonical name selection ---
    # For groups of 1, use the agency name directly.
    # For groups, find the best brand-level name.
    all_names = [r["raw_agency_name"] for r in rows]
    if len(rows) == 1:
        canonical_name = rows[0]["raw_agency_name"]
    else:
        canonical_name = canonical_brand_name(all_names)
        # Sanity: if stripped name is empty or single char, fall back to primary
        if not canonical_name or len(canonical_name) < 2:
            canonical_name = rows_sorted[0]["raw_agency_name"]

    # Use primary row for website, city, source
    primary = rows_sorted[0]

    # Collect all aliases (names that differ from canonical)
    aliases = sorted({n for n in all_names if n.lower() != canonical_name.lower()})

    # Collect ALL unique cities from all merged records
    all_cities = sorted({
        r["city_clean"] for r in rows
        if r["city_clean"] and r["city_clean"].lower() not in ("unknown", "")
    })
    primary_city = (
        primary["city_clean"]
        if primary["city_clean"].lower() not in ("unknown", "")
        else (all_cities[0] if all_cities else "unknown")
    )

    # Pick best housing/transport/job_types
    housing   = pick_enum_best(*[r["raw_housing"]  for r in rows_sorted])
    transport = pick_enum_best(*[r["raw_transport"] for r in rows_sorted])
    job_types = merge_job_types(*[r["raw_job_types"] for r in rows_sorted])
    salary    = pick_best(*[r["raw_salary_range"] for r in rows_sorted])
    notes     = merge_notes(*[r["raw_notes"] for r in rows_sorted])
    source_url = pick_best(*[r["raw_source"] for r in rows_sorted])
    website   = primary["raw_website"]

    # Determine if housing/transport came from notes detection
    h_from_notes = any(r.get("housing_from_notes") for r in rows)
    t_from_notes = any(r.get("transport_from_notes") for r in rows)

    housing_status   = detect_verification_status(notes, housing,   h_from_notes)
    transport_status = detect_verification_status(notes, transport,  t_from_notes)
    src_type         = detect_source_type(source_url)

    web_pages = build_web_pages(website, housing)

    # Housing type from notes
    housing_type = "UNKNOWN"
    if housing == "YES":
        nl = (notes or "").lower()
        if any(kw in nl for kw in ["shared", "gedeeld", "shared house"]):
            housing_type = "SHARED_HOUSE"
        elif any(kw in nl for kw in ["private", "eigen kamer", "private room"]):
            housing_type = "PRIVATE_ROOM"
        elif any(kw in nl for kw in ["studio", "appartement", "apartment"]):
            housing_type = "APARTMENT"
        else:
            housing_type = "SHARED_HOUSE"   # default for agency housing

    cid = f"agency_{canon_id:03d}"
    canon_id += 1

    record = {
        "canonical_id":         cid,
        "agency_name":          canonical_name,
        "slug":                 slugify(canonical_name),
        "website":              normalize_website(website),
        "web_pages":            web_pages,
        "city":                 primary_city,
        "known_cities":         all_cities,
        "housing":              housing,
        "housing_type":         housing_type,
        "transport":            transport,
        "job_types":            job_types,
        "salary_range":         salary if salary != "unknown" else None,
        "description":          (notes[:500] if notes else None),
        "housing_verification": {
            "value":       housing,
            "status":      housing_status,
            "source_url":  source_url,
            "source_type": src_type,
        },
        "transport_verification": {
            "value":       transport,
            "status":      transport_status,
            "source_url":  source_url,
            "source_type": src_type,
        },
        "source_url":         source_url,
        "source_type":        src_type,
        "last_checked_at":    TODAY,
        "aliases":            aliases,
        "original_names":     all_names,
        "duplicate_count":    len(rows),
        "confidence_score":   100 if len(rows) == 1 else 90,
        "merge_reason":       merge_reason,
        "raw_ids":            sorted(r["raw_id"] for r in rows),
    }

    for r in rows:
        dup_map_entries.append({
            "duplicate_row_id":  r["raw_id"],
            "duplicate_name":    r["raw_agency_name"],
            "canonical_id":      cid,
            "canonical_name":    canonical_name,
            "match_reason":      merge_reason,
            "is_canonical":      r["raw_agency_name"].lower() == canonical_name.lower(),
        })
        processed.add(r["raw_id"])

    return record

# Process domain / fuzzy groups
for g, info in merge_groups.items():
    rows = [staging[i] for i in sorted(info["ids"])]
    canonical_agencies.append(make_canonical(rows, info.get("reason", "unknown")))

# Singletons
for r in staging:
    if r["raw_id"] not in processed:
        canonical_agencies.append(make_canonical([r], "singleton"))

# Sort alphabetically by name
canonical_agencies.sort(key=lambda a: a["agency_name"].lower())

# Ensure unique slugs (append city abbreviation if collision)
seen_slugs: dict = {}
for a in canonical_agencies:
    base = a["slug"]
    if base in seen_slugs:
        # Disambiguate using primary city
        city_suffix = slugify(a["city"])[:10] if a["city"] != "unknown" else str(len(seen_slugs))
        a["slug"] = f"{base}-{city_suffix}"
    seen_slugs[a["slug"]] = a["canonical_id"]

print(f"\n[Phase 3B] Canonical agencies: {len(canonical_agencies)}")
print(f"  Duplicates removed: {len(staging) - len(canonical_agencies)}")
print(f"  Agencies with housing YES: {sum(1 for a in canonical_agencies if a['housing']=='YES')}")
print(f"  Agencies with transport YES: {sum(1 for a in canonical_agencies if a['transport']=='YES')}")
print(f"  Multi-city agencies: {sum(1 for a in canonical_agencies if len(a['known_cities'])>1)}")
print(f"  With job types: {sum(1 for a in canonical_agencies if a['job_types'])}")

# ─────────────────────────────────────────────────────────────────────────────
# PHASE 4 — WRITE OUTPUTS
# ─────────────────────────────────────────────────────────────────────────────

(OUT_DIR / "03_canonical_agencies.json").write_text(
    json.dumps(canonical_agencies, indent=2, ensure_ascii=False))
(OUT_DIR / "04_duplicate_map.json").write_text(
    json.dumps(dup_map_entries, indent=2, ensure_ascii=False))

branch_cities = {
    a["agency_name"]: a["known_cities"]
    for a in canonical_agencies if len(a["known_cities"]) > 1
}
(OUT_DIR / "05_branch_cities.json").write_text(
    json.dumps(branch_cities, indent=2, ensure_ascii=False))

alias_list = {
    a["agency_name"]: a["aliases"]
    for a in canonical_agencies if a["aliases"]
}
(OUT_DIR / "06_aliases.json").write_text(
    json.dumps(alias_list, indent=2, ensure_ascii=False))

review_flags = [p for p in fuzzy_all if not p["auto_merge"]]
(OUT_DIR / "07_review_flags.json").write_text(
    json.dumps(review_flags, indent=2, ensure_ascii=False))

# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

housing_yes     = sum(1 for a in canonical_agencies if a["housing"] == "YES")
housing_unk     = sum(1 for a in canonical_agencies if a["housing"] == "UNKNOWN")
transport_yes   = sum(1 for a in canonical_agencies if a["transport"] == "YES")
with_job_types  = sum(1 for a in canonical_agencies if a["job_types"])
with_multi_city = sum(1 for a in canonical_agencies if len(a["known_cities"]) > 1)
aliases_total   = sum(len(a["aliases"]) for a in canonical_agencies)

summary = {
    "generated_at":               TODAY,
    "raw_rows_imported":          len(staging),
    "duplicates_removed":         len(staging) - len(canonical_agencies),
    "unique_canonical_agencies":  len(canonical_agencies),
    "domain_merge_groups":        len(domain_dupes),
    "fuzzy_auto_merged":          len(fuzzy_auto_merge),
    "pairs_flagged_for_review":   len(review_flags),
    "agencies_housing_yes":       housing_yes,
    "agencies_housing_unknown":   housing_unk,
    "agencies_transport_yes":     transport_yes,
    "agencies_with_job_types":    with_job_types,
    "agencies_with_multi_cities": with_multi_city,
    "total_aliases":              aliases_total,
}
(OUT_DIR / "00_summary.json").write_text(
    json.dumps(summary, indent=2, ensure_ascii=False))

print("\n════════════════════════════════")
print("  PIPELINE COMPLETE — SUMMARY  ")
print("════════════════════════════════")
for k, v in summary.items():
    print(f"  {k}: {v}")
print(f"\n✓ All outputs written to {OUT_DIR}")

# Spot-checks
print("\n── Spot checks ──")
for name_fragment in ["Driessen", "OTTO Work Force", "Flexible Human Services", "Level One"]:
    found = [a for a in canonical_agencies if name_fragment.lower() in a["agency_name"].lower()]
    if found:
        a = found[0]
        print(f"  {name_fragment!r} → canonical: {a['agency_name']!r}  dup_count={a['duplicate_count']}  aliases={a['aliases']}")
    else:
        # Check aliases
        found = [a for a in canonical_agencies if any(name_fragment.lower() in al.lower() for al in a['aliases'])]
        if found:
            a = found[0]
            print(f"  {name_fragment!r} → IN ALIASES of {a['agency_name']!r}")
        else:
            print(f"  {name_fragment!r} → NOT FOUND")

print(f"\n  Housing YES agencies:")
for a in canonical_agencies:
    if a["housing"] == "YES":
        print(f"    {a['canonical_id']} | {a['agency_name']!r} | {a['city']} | {a['housing_verification']['status']}")
